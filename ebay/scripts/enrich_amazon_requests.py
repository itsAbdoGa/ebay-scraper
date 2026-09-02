import os
import random
import re
import shutil
import time
from datetime import date, datetime, timedelta
from http.cookiejar import CookieJar
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import (
    HTTPCookieProcessor,
    OpenerDirector,
    Request,
    build_opener,
)

from openpyxl import load_workbook
from selectolax.parser import HTMLParser

from lib.paths import COMBINED_XLSX, log_workbook_stop


DEFAULT_OUTPUT = COMBINED_XLSX.with_name(f"{COMBINED_XLSX.stem}_requests.xlsx")
AMAZON_URL_HEADERS = ("AMAZON URL", "URL: Amazon")
ENRICHMENT_DATE_HEADER = "LAST ENRICHMENT DATE"
ENRICHMENT_SKIP_DAYS = 3
OUTPUT_HEADERS = ("SALES RANK", "UPC", "BUYBOX", ENRICHMENT_DATE_HEADER)
OUTPUT_HEADER_ALIASES = {
    "BUYBOX": ("BUYBOX", "Buybox (30 days)"),
    ENRICHMENT_DATE_HEADER: (
        ENRICHMENT_DATE_HEADER,
        "Last Enrichment Date",
    ),
}
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/139.0.0.0 Safari/537.36"
)
BOT_MARKERS = (
    "<title>robot check</title>",
    "enter the characters you see below",
    "sorry, we just need to make sure you're not a robot",
    "click the button below to continue shopping",
    'id="captchacharacters"',
)


class AmazonBlockedError(RuntimeError):
    pass


def create_amazon_opener() -> OpenerDirector:
    return build_opener(HTTPCookieProcessor(CookieJar()))


def normalize_header(value: object) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def find_column(worksheet, candidates: tuple[str, ...]) -> int:
    expected = {normalize_header(candidate) for candidate in candidates}
    for cell in worksheet[1]:
        if normalize_header(cell.value) in expected:
            return cell.column
    raise ValueError(f"Missing required column; expected one of: {candidates}")


def ensure_output_columns(worksheet) -> dict[str, int]:
    existing = {
        normalize_header(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    columns: dict[str, int] = {}

    for header in OUTPUT_HEADERS:
        aliases = OUTPUT_HEADER_ALIASES.get(header, (header,))
        column = next(
            (
                existing[normalize_header(alias)]
                for alias in aliases
                if normalize_header(alias) in existing
            ),
            None,
        )
        if column is None:
            column = worksheet.max_column + 1
            worksheet.cell(row=1, column=column, value=header)
            existing[normalize_header(header)] = column
        columns[header] = column

    return columns


def parse_enrichment_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def is_recently_enriched(
    value: object,
    *,
    skip_days: int = ENRICHMENT_SKIP_DAYS,
    today: date | None = None,
) -> bool:
    if skip_days < 1:
        return False
    enriched_on = parse_enrichment_date(value)
    if enriched_on is None:
        return False
    reference = today or date.today()
    return reference - enriched_on < timedelta(days=skip_days)


def write_enrichment_date(
    worksheet,
    row_number: int,
    column: int,
    enriched_on: date | None = None,
) -> None:
    cell = worksheet.cell(
        row=row_number,
        column=column,
        value=enriched_on or date.today(),
    )
    cell.number_format = "YYYY-MM-DD"


def _has_cell_value(value: object) -> bool:
    return value is not None and str(value).strip() != ""


def _cell_values_equal(left: object, right: object) -> bool:
    if isinstance(left, datetime) or isinstance(right, datetime):
        return parse_enrichment_date(left) == parse_enrichment_date(right)
    if isinstance(left, date) or isinstance(right, date):
        return parse_enrichment_date(left) == parse_enrichment_date(right)
    return str(left or "").strip() == str(right or "").strip()


def _copy_filled_cell(source_cell, dest_cell, *, number_format: str | None = None) -> bool:
    if not _has_cell_value(source_cell.value):
        return False
    if _cell_values_equal(source_cell.value, dest_cell.value):
        return False
    dest_cell.value = source_cell.value
    dest_cell.number_format = (
        number_format or source_cell.number_format or dest_cell.number_format
    )
    return True


def sync_enrichment_to_scrape_workbook(
    source,
    scrape_path: Path,
    *,
    start_row: int = 2,
    end_row: int | None = None,
) -> tuple[int, int]:
    """Copy enrichment columns into the scrape workbook. Returns (updated, mismatched)."""
    scrape_path = scrape_path.resolve()
    source_workbook = None
    if hasattr(source, "cell"):
        source_worksheet = source
    else:
        source_path = Path(source).resolve()
        if source_path == scrape_path:
            return 0, 0
        source_workbook = load_workbook(source_path)
        source_worksheet = source_workbook.active

    dest_workbook = None
    try:
        if not scrape_path.exists():
            raise FileNotFoundError(f"Scrape workbook does not exist: {scrape_path}")

        dest_workbook = load_workbook(scrape_path)
        dest_worksheet = dest_workbook.active
        source_columns = ensure_output_columns(source_worksheet)
        dest_columns = ensure_output_columns(dest_worksheet)
        source_asin = find_column(source_worksheet, ("ASIN",))
        dest_asin = find_column(dest_worksheet, ("ASIN",))
        last_row = min(
            source_worksheet.max_row,
            dest_worksheet.max_row,
            end_row or min(source_worksheet.max_row, dest_worksheet.max_row),
        )
        updated = 0
        mismatched = 0
        for row_number in range(start_row, last_row + 1):
            source_asin_value = source_worksheet.cell(
                row=row_number, column=source_asin
            ).value
            dest_asin_value = dest_worksheet.cell(
                row=row_number, column=dest_asin
            ).value
            if str(source_asin_value or "").strip().casefold() != str(
                dest_asin_value or ""
            ).strip().casefold():
                mismatched += 1
                continue

            changed = False
            for header, number_format in (
                ("SALES RANK", "@"),
                ("UPC", "@"),
                ("BUYBOX", "$0.00"),
            ):
                changed = (
                    _copy_filled_cell(
                        source_worksheet.cell(
                            row=row_number, column=source_columns[header]
                        ),
                        dest_worksheet.cell(
                            row=row_number, column=dest_columns[header]
                        ),
                        number_format=number_format,
                    )
                    or changed
                )

            source_date_cell = source_worksheet.cell(
                row=row_number,
                column=source_columns[ENRICHMENT_DATE_HEADER],
            )
            dest_date_cell = dest_worksheet.cell(
                row=row_number,
                column=dest_columns[ENRICHMENT_DATE_HEADER],
            )
            if _has_cell_value(source_date_cell.value):
                changed = (
                    _copy_filled_cell(
                        source_date_cell,
                        dest_date_cell,
                        number_format="YYYY-MM-DD",
                    )
                    or changed
                )
            elif changed:
                write_enrichment_date(
                    dest_worksheet,
                    row_number,
                    dest_columns[ENRICHMENT_DATE_HEADER],
                )

            if changed:
                updated += 1

        atomic_save(dest_workbook, scrape_path)
        return updated, mismatched
    finally:
        if dest_workbook is not None:
            dest_workbook.close()
        if source_workbook is not None:
            source_workbook.close()


def persist_enrichment_results(
    workbook,
    worksheet,
    *,
    output_path: Path,
    scrape_path: Path,
    start_row: int,
    end_row: int | None,
) -> None:
    try:
        atomic_save(workbook, output_path)
        print(f"Saved working copy: {output_path}")
    except Exception as error:
        print(f"Could not save working copy: {type(error).__name__}: {error}")

    if end_row is None or end_row < start_row:
        return
    if scrape_path.resolve() == output_path.resolve():
        return
    try:
        updated, mismatched = sync_enrichment_to_scrape_workbook(
            worksheet,
            scrape_path,
            start_row=start_row,
            end_row=end_row,
        )
        print(f"Updated scrape workbook {scrape_path}: {updated} rows")
        if mismatched:
            print(f"  Skipped {mismatched} rows whose ASIN did not match")
    except Exception as error:
        print(f"Could not update scrape workbook: {type(error).__name__}: {error}")


def fetch_amazon_html(
    url: str,
    *,
    timeout_seconds: float,
    retries: int,
    opener: OpenerDirector | None = None,
) -> str:
    client = opener or create_amazon_opener()
    request = Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": (
                "text/html,application/xhtml+xml,application/xml;q=0.9,"
                "image/avif,image/webp,*/*;q=0.8"
            ),
            "Accept-Language": "en-US,en;q=0.9",
            "DNT": "1",
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
        },
    )

    last_error: Exception | None = None
    for attempt in range(retries + 1):
        try:
            with client.open(request, timeout=timeout_seconds) as response:
                charset = response.headers.get_content_charset() or "utf-8"
                html = response.read().decode(charset, errors="replace")
            lowered = html.casefold()
            if any(marker in lowered for marker in BOT_MARKERS):
                raise AmazonBlockedError("Amazon returned a bot-check page")
            return html
        except AmazonBlockedError:
            raise
        except HTTPError as error:
            if error.code in (429, 503):
                raise AmazonBlockedError(
                    f"Amazon returned blocking status HTTP {error.code}"
                ) from error
            last_error = error
        except (URLError, TimeoutError) as error:
            last_error = error

        if attempt < retries:
            backoff = min(30.0, (2**attempt) + random.uniform(0.5, 1.5))
            time.sleep(backoff)

    assert last_error is not None
    raise last_error


def parse_price_text(value: str) -> float | str:
    match = re.search(r"\$?\s*([\d,]+(?:\.\d{1,2})?)", value)
    if match is None:
        return ""
    return float(match.group(1).replace(",", ""))


def extract_buybox_price(tree: HTMLParser) -> float | str:
    accessibility_price = tree.css_first(
        "#apex-pricetopay-accessibility-label"
    )
    if accessibility_price is not None:
        parsed = parse_price_text(accessibility_price.text(strip=True))
        if parsed != "":
            return parsed

    hidden_amount = tree.css_first(
        'input[name="items[0.base][customerVisiblePrice][amount]"]'
    )
    if hidden_amount is not None:
        parsed = parse_price_text(hidden_amount.attributes.get("value", ""))
        if parsed != "":
            return parsed

    hidden_display_price = tree.css_first(
        'input[name="items[0.base][customerVisiblePrice][displayString]"]'
    )
    if hidden_display_price is not None:
        parsed = parse_price_text(
            hidden_display_price.attributes.get("value", "")
        )
        if parsed != "":
            return parsed

    selectors = (
        ".apex-core-price-identifier .a-price.apex-pricetopay-value",
        ".price-line-wrapper .a-price.priceToPay",
        ".price-line-wrapper .a-price.apex-pricetopay-value",
        "#corePrice_feature_div .a-price.priceToPay",
    )
    price = next((tree.css_first(selector) for selector in selectors if tree.css_first(selector)), None)
    if price is None:
        return ""

    offscreen = price.css_first(".a-offscreen")
    if offscreen is not None:
        parsed = parse_price_text(offscreen.text(strip=True))
        if parsed != "":
            return parsed

    whole = price.css_first(".a-price-whole")
    fraction = price.css_first(".a-price-fraction")
    whole_digits = re.sub(r"\D", "", whole.text(strip=True) if whole else "")
    fraction_digits = re.sub(
        r"\D",
        "",
        fraction.text(strip=True) if fraction else "",
    )
    if not whole_digits:
        return ""
    amount = whole_digits
    if fraction_digits:
        amount += f".{fraction_digits[:2]}"
    return float(amount)


def extract_product_details(html: str) -> dict[str, str | float]:
    tree = HTMLParser(html)
    sales_rank = ""
    upc = ""
    gtin = ""

    for row in tree.css("table.prodDetTable tr"):
        heading = row.css_first("th")
        value = row.css_first("td")
        if heading is None or value is None:
            continue

        label = " ".join(heading.text(separator=" ", strip=True).split()).casefold()
        value_text = " ".join(value.text(separator=" ", strip=True).split())
        if label == "upc":
            match = re.search(r"\b\d{8,14}\b", value_text)
            upc = match.group(0) if match else ""
        elif label == "global trade identification number":
            match = re.search(r"\b\d{8,14}\b", value_text)
            gtin = match.group(0) if match else ""
        elif label == "best sellers rank":
            match = re.search(r"#\s*([\d,]+)", value_text)
            sales_rank = match.group(1) if match else ""

    return {
        "SALES RANK": sales_rank,
        "UPC": upc or gtin,
        "BUYBOX": extract_buybox_price(tree),
    }


def atomic_save(workbook, output_path: Path) -> None:
    temporary_path = output_path.with_name(
        f".{output_path.stem}.tmp{output_path.suffix}"
    )
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, output_path)
    finally:
        try:
            temporary_path.unlink(missing_ok=True)
        except OSError:
            pass


def enrich_workbook(
    *,
    input_path: Path = COMBINED_XLSX,
    output_path: Path = DEFAULT_OUTPUT,
    timeout_seconds: float = 10.0,
    retries: int = 2,
    delay_seconds: float = 1.5,
    save_every: int = 10,
    max_consecutive_blocks: int = 3,
    skip_enriched: bool = False,
    skip_recent_days: int = ENRICHMENT_SKIP_DAYS,
    fresh_copy: bool = False,
    start: int | None = None,
    end: int | None = None,
) -> int:
    input_path = input_path.resolve()
    output_path = output_path.resolve()

    if input_path == output_path:
        print("Input and output must differ; the original workbook is never modified.")
        return 1
    if not input_path.exists():
        print(f"Input workbook does not exist: {input_path}")
        return 1
    if (
        timeout_seconds <= 0
        or retries < 0
        or delay_seconds < 0
        or save_every < 1
        or max_consecutive_blocks < 1
        or skip_recent_days < 0
    ):
        print(
            "Timeout/save/block limits must be positive; "
            "retries/delay/skip-recent-days cannot be negative."
        )
        return 1

    if fresh_copy or not output_path.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(input_path, output_path)
        print(f"Created working copy: {output_path}")
    else:
        print(f"Resuming existing working copy: {output_path}")

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    try:
        url_column = find_column(worksheet, AMAZON_URL_HEADERS)
        output_columns = ensure_output_columns(worksheet)
    except ValueError as error:
        workbook.close()
        print(error)
        return 1

    first_row = max(2, start or 2)
    last_row = min(worksheet.max_row, end or worksheet.max_row)
    rows = range(first_row, last_row + 1)
    completed = 0
    failures = 0
    unsaved = 0
    consecutive_blocks = 0
    last_processed_row = first_row - 1
    stop_reason: str | None = None
    opener = create_amazon_opener()

    print(f"Rows: {first_row}-{last_row}")
    if skip_recent_days:
        print(f"Skipping rows enriched in the last {skip_recent_days} days")
    try:
        for position, row_number in enumerate(rows, start=1):
            last_processed_row = row_number
            amazon_url = str(
                worksheet.cell(row=row_number, column=url_column).value or ""
            ).strip()
            if not amazon_url:
                print(f"[{position}] Row {row_number}: no Amazon URL")
                continue

            last_enriched = worksheet.cell(
                row=row_number,
                column=output_columns[ENRICHMENT_DATE_HEADER],
            ).value
            if is_recently_enriched(last_enriched, skip_days=skip_recent_days):
                print(
                    f"[{position}] Row {row_number}: skipped "
                    f"(enriched in the last {skip_recent_days} days)"
                )
                continue

            rank_cell = worksheet.cell(
                row=row_number,
                column=output_columns["SALES RANK"],
            )
            upc_cell = worksheet.cell(row=row_number, column=output_columns["UPC"])
            if (
                skip_enriched
                and str(rank_cell.value or "").strip()
                and str(upc_cell.value or "").strip()
            ):
                print(f"[{position}] Row {row_number}: already enriched")
                continue

            print(f"[{position}] Row {row_number}: {amazon_url}")
            try:
                html = fetch_amazon_html(
                    amazon_url,
                    timeout_seconds=timeout_seconds,
                    retries=retries,
                    opener=opener,
                )
                details = extract_product_details(html)
                consecutive_blocks = 0

                for header in ("SALES RANK", "UPC"):
                    value = details[header]
                    if value == "":
                        continue
                    cell = worksheet.cell(
                        row=row_number,
                        column=output_columns[header],
                        value=value,
                    )
                    cell.number_format = "@"

                buybox = details["BUYBOX"]
                if buybox != "":
                    buybox_cell = worksheet.cell(
                        row=row_number,
                        column=output_columns["BUYBOX"],
                        value=buybox,
                    )
                    buybox_cell.number_format = "$0.00"

                write_enrichment_date(
                    worksheet,
                    row_number,
                    output_columns[ENRICHMENT_DATE_HEADER],
                )
                completed += 1
                unsaved += 1
                print(
                    "  "
                    f"rank={details['SALES RANK'] or '-'}, "
                    f"upc={details['UPC'] or '-'}, "
                    f"buybox={details['BUYBOX'] or '-'}"
                )
            except AmazonBlockedError as error:
                failures += 1
                consecutive_blocks += 1
                print(f"  Blocked: {error}")
                if consecutive_blocks >= max_consecutive_blocks:
                    print(
                        f"Stopping after {consecutive_blocks} consecutive "
                        "bot checks; completed rows will be saved."
                    )
                    stop_reason = (
                        f"{consecutive_blocks} consecutive Amazon bot checks"
                    )
                    break
                cooldown = min(60.0, 10.0 * (2 ** (consecutive_blocks - 1)))
                print(f"  Cooling down for {cooldown:g} seconds.")
                time.sleep(cooldown)
            except Exception as error:
                failures += 1
                consecutive_blocks = 0
                print(f"  Failed: {type(error).__name__}: {error}")

            if unsaved >= save_every:
                atomic_save(workbook, output_path)
                unsaved = 0
            if delay_seconds:
                time.sleep(
                    random.uniform(
                        delay_seconds * 0.75,
                        delay_seconds * 1.25,
                    )
                )
    except KeyboardInterrupt:
        print("Stopped by user; saving completed rows.")
        stop_reason = "stopped by user"
    except Exception as error:
        print(f"Enrichment stopped: {type(error).__name__}: {error}")
        stop_reason = f"{type(error).__name__}: {error}"
        failures += 1
    finally:
        if stop_reason:
            log_workbook_stop(
                workbook_path=output_path,
                excel_row=(
                    last_processed_row
                    if last_processed_row >= first_row
                    else first_row
                ),
                reason=stop_reason,
                script="enrich-amazon",
                extra={"scrape_file": str(input_path)},
            )
        persist_enrichment_results(
            workbook,
            worksheet,
            output_path=output_path,
            scrape_path=input_path,
            start_row=first_row,
            end_row=last_processed_row,
        )
        try:
            workbook.close()
        except Exception:
            pass

    print(f"Enriched: {completed}; failed: {failures}")
    return 0 if failures == 0 else 2
