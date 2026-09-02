import random
import shutil
import time
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import (
    Page,
    TimeoutError as PlaywrightTimeoutError,
    sync_playwright,
)

from lib.paths import COMBINED_XLSX, log_workbook_stop
from scripts.enrich_amazon_requests import (
    AMAZON_URL_HEADERS,
    BOT_MARKERS,
    DEFAULT_OUTPUT,
    ENRICHMENT_DATE_HEADER,
    ENRICHMENT_SKIP_DAYS,
    AmazonBlockedError,
    atomic_save,
    ensure_output_columns,
    extract_product_details,
    find_column,
    is_recently_enriched,
    parse_price_text,
    persist_enrichment_results,
    write_enrichment_date,
)


DEFAULT_PROFILE_DIR = (
    Path(__file__).resolve().parent.parent
    / "data"
    / "browser"
    / "amazon_profile"
)
BROWSER_PROFILES = (
    {
        "user_agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/139.0.0.0 Safari/537.36"
        ),
        "sec_ch_ua": '"Chromium";v="139", "Not_A Brand";v="24"',
        "accept_language": "en-US,en;q=0.9",
    },
    {
        "user_agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/138.0.0.0 Safari/537.36"
        ),
        "sec_ch_ua": '"Chromium";v="138", "Not_A Brand";v="24"',
        "accept_language": "en-US,en;q=0.8",
    },
    {
        "user_agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/137.0.0.0 Safari/537.36"
        ),
        "sec_ch_ua": '"Chromium";v="137", "Not_A Brand";v="24"',
        "accept_language": "en-US,en;q=0.9",
    },
)
VIEWPORTS = (
    {"width": 1366, "height": 768},
    {"width": 1440, "height": 900},
    {"width": 1536, "height": 864},
    {"width": 1920, "height": 1080},
)
BROWSER_RESTART_EVERY = 1000
BROWSER_RESTART_PAUSE_SECONDS = 1.5


class AmazonDetailsMissingError(RuntimeError):
    pass


def fetch_amazon_html(
    page: Page,
    url: str,
    *,
    timeout_seconds: float,
    retries: int,
) -> str:
    last_error: Exception | None = None

    for attempt in range(retries + 1):
        try:
            response = page.goto(
                url,
                wait_until="domcontentloaded",
                timeout=int(timeout_seconds * 1000),
            )
            if response is not None and response.status in (429, 503):
                raise AmazonBlockedError(
                    f"Amazon returned blocking status HTTP {response.status}"
                )

            # One scroll is enough to trigger Amazon's lazy-loaded product details.
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            element_timeout_ms = random.uniform(4_000, 6_000)
            try:
                page.wait_for_selector(
                    "table.prodDetTable",
                    state="attached",
                    timeout=element_timeout_ms,
                )
            except PlaywrightTimeoutError as error:
                raise AmazonDetailsMissingError(
                    "product details table did not load within "
                    f"{element_timeout_ms / 1000:.1f} seconds"
                ) from error

            html = page.content()
            lowered = html.casefold()
            if any(marker in lowered for marker in BOT_MARKERS):
                raise AmazonBlockedError("Amazon returned a bot-check page")
            return html
        except AmazonBlockedError:
            raise
        except AmazonDetailsMissingError:
            raise
        except PlaywrightTimeoutError as error:
            last_error = error

        if attempt < retries:
            time.sleep(min(30.0, (2**attempt) + random.uniform(0.5, 1.5)))

    assert last_error is not None
    raise last_error


def launch_amazon_browser(playwright, *, user_data_dir: Path, headless: bool):
    browser_profile = random.choice(BROWSER_PROFILES)
    viewport = random.choice(VIEWPORTS)
    user_data_dir.mkdir(parents=True, exist_ok=True)
    context = playwright.chromium.launch_persistent_context(
        user_data_dir=str(user_data_dir),
        headless=headless,
        user_agent=browser_profile["user_agent"],
        locale="en-US",
        timezone_id="America/New_York",
        viewport=viewport,
        color_scheme="light",
        extra_http_headers={
            "Accept-Language": browser_profile["accept_language"],
            "Sec-CH-UA": browser_profile["sec_ch_ua"],
            "Sec-CH-UA-Mobile": "?0",
            "Sec-CH-UA-Platform": '"Windows"',
            "Upgrade-Insecure-Requests": "1",
        },
    )
    context.add_init_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    page = context.pages[0] if context.pages else context.new_page()
    return context, page


def close_amazon_browser(context) -> None:
    try:
        context.close()
    except Exception as error:
        print(f"Browser close failed: {type(error).__name__}: {error}")


def enrich_workbook(
    *,
    input_path: Path = COMBINED_XLSX,
    output_path: Path = DEFAULT_OUTPUT,
    timeout_seconds: float = 10.0,
    retries: int = 2,
    delay_seconds: float = 0.0,
    min_buybox_price: float = 40.0,
    save_every: int = 10,
    max_consecutive_blocks: int = 3,
    fresh_copy: bool = False,
    user_data_dir: Path = DEFAULT_PROFILE_DIR,
    headless: bool = False,
    start: int | None = None,
    end: int | None = None,
    skip_recent_days: int = ENRICHMENT_SKIP_DAYS,
    browser_restart_every: int = BROWSER_RESTART_EVERY,
) -> int:
    input_path = input_path.resolve()
    output_path = output_path.resolve()
    user_data_dir = user_data_dir.resolve()

    if input_path == output_path:
        print("Input and output must differ; the original workbook is never modified.")
        return 1
    if not input_path.exists():
        print(f"Input workbook does not exist: {input_path}")
        return 1
    if (
        timeout_seconds <= 0
        or retries < 0
        or delay_seconds < 0
        or min_buybox_price < 0
        or save_every < 1
        or max_consecutive_blocks < 1
        or skip_recent_days < 0
        or browser_restart_every < 0
    ):
        print(
            "Timeout/save/block limits must be positive; "
            "retries/delay/minimum buybox/skip-recent-days/"
            "browser-restart-every cannot be negative."
        )
        return 1

    if fresh_copy or not output_path.exists():
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(input_path, output_path)
        print(f"Created working copy: {output_path}")
    else:
        print(f"Resuming existing working copy: {output_path}")

    workbook = load_workbook(output_path)
    worksheet = workbook.active
    try:
        url_column = find_column(worksheet, AMAZON_URL_HEADERS)
        output_columns = ensure_output_columns(worksheet)
    except ValueError as error:
        workbook.close()
        print(error)
        return 1

    first_row = max(2, start or 2)
    last_row = min(worksheet.max_row, end or worksheet.max_row)
    rows = range(first_row, last_row + 1)
    completed = 0
    skipped = 0
    failures = 0
    unsaved = 0
    consecutive_blocks = 0
    last_processed_row = first_row - 1
    stop_reason: str | None = None

    print(f"Rows: {first_row}-{last_row}")
    print(f"Persistent browser profile: {user_data_dir}")
    if skip_recent_days:
        print(f"Skipping rows enriched in the last {skip_recent_days} days")
    if browser_restart_every:
        print(f"Restarting browser every {browser_restart_every} rows")
    context = None
    try:
        with sync_playwright() as playwright:
            context, page = launch_amazon_browser(
                playwright,
                user_data_dir=user_data_dir,
                headless=headless,
            )
            rows_since_browser_start = 0
            try:
                for position, row_number in enumerate(rows, start=1):
                    last_processed_row = row_number
                    if (
                        browser_restart_every
                        and rows_since_browser_start >= browser_restart_every
                    ):
                        print(
                            f"Restarting browser after {position - 1} rows "
                            "to free memory"
                        )
                        if unsaved:
                            atomic_save(workbook, output_path)
                            unsaved = 0
                        close_amazon_browser(context)
                        context = None
                        time.sleep(BROWSER_RESTART_PAUSE_SECONDS)
                        context, page = launch_amazon_browser(
                            playwright,
                            user_data_dir=user_data_dir,
                            headless=headless,
                        )
                        rows_since_browser_start = 0

                    rows_since_browser_start += 1
                    amazon_url = str(
                        worksheet.cell(
                            row=row_number,
                            column=url_column,
                        ).value
                        or ""
                    ).strip()
                    if not amazon_url:
                        print(f"[{position}] Row {row_number}: no Amazon URL")
                        continue

                    last_enriched = worksheet.cell(
                        row=row_number,
                        column=output_columns[ENRICHMENT_DATE_HEADER],
                    ).value
                    if is_recently_enriched(
                        last_enriched,
                        skip_days=skip_recent_days,
                    ):
                        skipped += 1
                        print(
                            f"[{position}] Row {row_number}: skipped "
                            f"(enriched in the last {skip_recent_days} days)"
                        )
                        continue

                    buybox_value = worksheet.cell(
                        row=row_number,
                        column=output_columns["BUYBOX"],
                    ).value
                    buybox_price = parse_price_text(str(buybox_value or ""))
                    has_buybox = isinstance(buybox_price, float)
                    if has_buybox and buybox_price <= min_buybox_price:
                        skipped += 1
                        print(
                            f"[{position}] Row {row_number}: skipped "
                            f"(buybox must be over ${min_buybox_price:,.2f})"
                        )
                        continue

                    print(f"[{position}] Row {row_number}: {amazon_url}")
                    try:
                        html = fetch_amazon_html(
                            page,
                            amazon_url,
                            timeout_seconds=timeout_seconds,
                            retries=retries,
                        )
                        details = extract_product_details(html)
                        consecutive_blocks = 0
                        if (
                            not details["SALES RANK"]
                            and not details["UPC"]
                            and not details["BUYBOX"]
                        ):
                            skipped += 1
                            print(
                                "  Skipped: page contains no rank, UPC, or buybox"
                            )
                            continue

                        for header in ("SALES RANK", "UPC"):
                            value = details[header]
                            if value == "":
                                continue
                            cell = worksheet.cell(
                                row=row_number,
                                column=output_columns[header],
                                value=value,
                            )
                            cell.number_format = "@"

                        buybox = details["BUYBOX"]
                        if buybox != "":
                            buybox_cell = worksheet.cell(
                                row=row_number,
                                column=output_columns["BUYBOX"],
                                value=buybox,
                            )
                            buybox_cell.number_format = "$0.00"

                        write_enrichment_date(
                            worksheet,
                            row_number,
                            output_columns[ENRICHMENT_DATE_HEADER],
                        )
                        completed += 1
                        unsaved += 1
                        print(
                            "  "
                            f"rank={details['SALES RANK'] or '-'}, "
                            f"upc={details['UPC'] or '-'}, "
                            f"buybox={details['BUYBOX'] or '-'}"
                        )
                    except AmazonDetailsMissingError as error:
                        skipped += 1
                        consecutive_blocks = 0
                        print(f"  Skipped: {error}")
                    except AmazonBlockedError as error:
                        failures += 1
                        consecutive_blocks += 1
                        print(f"  Blocked: {error}")
                        if consecutive_blocks >= max_consecutive_blocks:
                            print(
                                f"Stopping after {consecutive_blocks} consecutive "
                                "bot checks; completed rows will be saved."
                            )
                            stop_reason = (
                                f"{consecutive_blocks} consecutive Amazon bot checks"
                            )
                            break
                        cooldown = min(
                            15.0,
                            3.0 * (2 ** (consecutive_blocks - 1)),
                        )
                        print(f"  Cooling down for {cooldown:g} seconds.")
                        time.sleep(cooldown)
                    except Exception as error:
                        failures += 1
                        consecutive_blocks = 0
                        print(f"  Failed: {type(error).__name__}: {error}")

                    if unsaved >= save_every:
                        atomic_save(workbook, output_path)
                        unsaved = 0
                    if delay_seconds:
                        time.sleep(
                            random.uniform(
                                delay_seconds * 0.75,
                                delay_seconds * 1.25,
                            )
                        )
            finally:
                if context is not None:
                    close_amazon_browser(context)
    except KeyboardInterrupt:
        print("Stopped by user; saving completed rows.")
        stop_reason = "stopped by user"
    except Exception as error:
        print(f"Enrichment stopped: {type(error).__name__}: {error}")
        stop_reason = f"{type(error).__name__}: {error}"
        failures += 1
    finally:
        if stop_reason:
            log_workbook_stop(
                workbook_path=output_path,
                excel_row=(
                    last_processed_row
                    if last_processed_row >= first_row
                    else first_row
                ),
                reason=stop_reason,
                script="enrich-amazon",
                extra={"scrape_file": str(input_path)},
            )
        persist_enrichment_results(
            workbook,
            worksheet,
            output_path=output_path,
            scrape_path=input_path,
            start_row=first_row,
            end_row=last_processed_row,
        )
        try:
            workbook.close()
        except Exception:
            pass

    print(f"Enriched: {completed}; skipped: {skipped}; failed: {failures}")
    return 0 if failures == 0 else 2
