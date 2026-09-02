import csv
import json
import os
import re
import sys
from collections.abc import Callable
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from lib.ebay_scraper import (
    BROWSER_RESTART_EVERY,
    EbayShipToNotUsError,
    browser_session,
    describe_ebay_cookie_session,
    evaluate_winning_listing,
    fetch_amazon_sales_rank,
    fill_missing_seller_reviews,
    filter_us_listings,
    format_below_buybox_listing_logs,
    format_block_log,
    format_cheapest_listing_log,
    load_ebay_cookies,
    parse_price,
    refresh_and_verify_ship_to_us,
    scrape_search_page,
    scrape_search_page_from_html,
)
from lib.paths import (
    COMBINED_XLSX,
    EBAY_COOKIES_FILE,
    IDENTIFIER_NO_MATCH_HISTORY_JSON,
    WINNING_LISTINGS_JSON,
    WINNING_LISTINGS_HISTORY_JSON,
    WINNING_LISTINGS_XLSX,
    create_workbook_backup,
    ensure_data_dirs,
    log_workbook_stop,
    prune_directory_workbook_backups,
)

TEXT_COLUMNS = {"EAN", "ASIN"}
WINNER_HEADERS = [
    "title",
    "ASIN",
    "EAN",
    "Brand",
    "BUYBOX",
    "AMAZON URL",
    "EBAY full cost",
    "SELLER",
    "SELLER REVIEWS",
    "LISTING DATE",
    "ROI",
    "EBAY listing URL",
    "ebay listing query",
]


@dataclass
class ScrapeSettings:
    """Tweak these in ebay/main.py, then run that file."""

    input_csv: Path = COMBINED_XLSX
    output_xlsx: Path = WINNING_LISTINGS_XLSX
    live_json: Path = WINNING_LISTINGS_JSON
    winner_history: Path = WINNING_LISTINGS_HISTORY_JSON
    identifier_no_match_history: Path = IDENTIFIER_NO_MATCH_HISTORY_JSON

    start_row: int | None = None
    end_row: int | None = None
    limit: int | None = None

    include_brands: tuple[str, ...] = ()
    exclude_brands: tuple[str, ...] = ()

    min_roi_percent: float = 80.0
    max_roi_percent: float | None = 300.0
    max_listing_age_days: int = 2
    min_seller_reviews: int = 50
    min_sales_rank: int | None = None
    max_sales_rank: int | None = 500_000
    min_drop_count: int | None = None
    min_buybox_price: float | None = 40.0

    upc_as_well: bool = True
    cleaned_title_as_well: bool = False
    titles_only: bool = False
    skip_previously_won: bool = True
    winner_history_retention_days: int = 3
    identifier_no_match_retention_days: int = 3

    html_path: Path | None = None
    cookies_file: Path | None = EBAY_COOKIES_FILE
    cookie_header: str | None = None

    def validate(self) -> None:
        if self.start_row is not None and self.start_row < 2:
            raise ValueError("start_row must be at least 2 (Excel row 1 is the header)")
        if self.end_row is not None and self.end_row < 2:
            raise ValueError("end_row must be at least 2 (Excel row 1 is the header)")
        if (
            self.start_row is not None
            and self.end_row is not None
            and self.start_row > self.end_row
        ):
            raise ValueError("start_row cannot be greater than end_row")
        if self.min_sales_rank is not None and self.min_sales_rank < 0:
            raise ValueError("min_sales_rank cannot be negative")
        if self.max_sales_rank is not None and self.max_sales_rank < 0:
            raise ValueError("max_sales_rank cannot be negative")
        if (
            self.min_sales_rank is not None
            and self.max_sales_rank is not None
            and self.min_sales_rank > self.max_sales_rank
        ):
            raise ValueError("min_sales_rank cannot exceed max_sales_rank")
        if self.min_drop_count is not None and self.min_drop_count < 0:
            raise ValueError("min_drop_count cannot be negative")
        if self.min_buybox_price is not None and self.min_buybox_price < 0:
            raise ValueError("min_buybox_price cannot be negative")
        if (
            self.max_roi_percent is not None
            and self.max_roi_percent < self.min_roi_percent
        ):
            raise ValueError("max_roi_percent cannot be less than min_roi_percent")
        if self.min_seller_reviews < 0:
            raise ValueError("min_seller_reviews cannot be negative")
        if (
            self.winner_history_retention_days is not None
            and self.winner_history_retention_days < 1
        ):
            raise ValueError("winner_history_retention_days must be at least 1")
        if self.identifier_no_match_retention_days < 1:
            raise ValueError("identifier_no_match_retention_days must be at least 1")

    def winner_kwargs(self) -> dict:
        return {
            "min_roi_percent": self.min_roi_percent,
            "max_roi_percent": self.max_roi_percent,
            "max_listing_age_days": self.max_listing_age_days,
            "min_seller_reviews": self.min_seller_reviews,
        }


def load_products(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "") for value in next(values, ())]
            rows = [dict(zip(headers, row)) for row in values]
        finally:
            workbook.close()
    else:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))

    products = []
    # Row 1 contains headers, so these numbers match Excel's visible row numbers.
    for source_row, row in enumerate(rows, start=2):
        products.append(
            {
                "csv_row": source_row,
                "title": str(row.get("TITLE") or ""),
                "asin": str(row.get("ASIN") or ""),
                "ean": str(row.get("EAN") or ""),
                "upc": str(row.get("UPC") or ""),
                "cleaned_title": str(row.get("CLEANED TITLE") or "").strip(),
                "sales_rank": str(row.get("SALES RANK") or ""),
                "drops_count": str(row.get("DROPS (90 DAYS)") or ""),
                "brand": str(row.get("Brand") or ""),
                "buybox_price": str(row.get("Buybox (30 days)") or ""),
                "amazon_url": str(row.get("URL: Amazon") or ""),
                "search_url": str(row.get("ebay listing") or "").strip(),
            }
        )
    return products


def parse_sales_rank(value: object) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return int(value)

    match = re.search(r"\d[\d,]*", str(value))
    return int(match.group(0).replace(",", "")) if match else None


def sales_rank_qualified(
    value: object,
    *,
    min_sales_rank: int | None,
    max_sales_rank: int | None,
) -> bool:
    sales_rank = parse_sales_rank(value)
    return (
        sales_rank is not None
        and (min_sales_rank is None or sales_rank >= min_sales_rank)
        and (max_sales_rank is None or sales_rank <= max_sales_rank)
    )


def ebay_search_url_for_identifier(search_url: str, identifier: str) -> str:
    parts = urlsplit(search_url)
    query = parse_qsl(parts.query, keep_blank_values=True)
    replaced = False
    updated_query: list[tuple[str, str]] = []
    for key, value in query:
        if key.casefold() in {"_nkw", "nkw"}:
            updated_query.append((key, identifier))
            replaced = True
        else:
            updated_query.append((key, value))
    if not replaced:
        updated_query.append(("_nkw", identifier))
    return urlunsplit(
        (
            parts.scheme,
            parts.netloc,
            parts.path,
            urlencode(updated_query),
            parts.fragment,
        )
    )


def identifier_query_key(identifier_type: str, identifier: object) -> str:
    value = str(identifier or "").strip().casefold()
    return f"{identifier_type.upper()}:{value}" if value else ""


def scrape_stop_extra(product: dict[str, str] | None) -> dict:
    if not product:
        return {}
    return {
        "query_type": str(product.get("search_identifier_type") or ""),
        "query": str(
            product.get("search_identifier")
            or product.get("ean")
            or product.get("asin")
            or ""
        ),
        "asin": str(product.get("asin") or ""),
        "selection_position": product.get("selection_position"),
        "selection_total": product.get("selection_total"),
    }


def expand_identifier_searches(
    products: list[dict[str, str]],
    *,
    upc_as_well: bool,
    cleaned_title_as_well: bool,
    titles_only: bool = False,
    skipped_identifier_keys: set[str] | None = None,
) -> list[dict[str, str]]:
    expanded: list[dict[str, str]] = []
    emitted_title_searches: set[tuple[str, str]] = set()
    skipped_identifier_keys = skipped_identifier_keys or set()
    for product in products:
        ean = product.get("ean", "").strip()
        upc = product.get("upc", "").strip()
        asin = product.get("asin", "").strip().casefold()
        cleaned_title = product.get("cleaned_title", "").strip()
        identifiers: list[tuple[str, str]] = []
        title_key = (asin, cleaned_title.casefold())
        include_title = (
            (titles_only or cleaned_title_as_well)
            and asin
            and cleaned_title
            and title_key not in emitted_title_searches
        )
        if titles_only:
            if include_title:
                identifiers.append(("TITLE", cleaned_title))
                emitted_title_searches.add(title_key)
            if not identifiers:
                continue
        else:
            has_identifier_candidate = bool(ean or (upc_as_well and upc and upc != ean))
            if ean and identifier_query_key("EAN", ean) not in skipped_identifier_keys:
                identifiers.append(("EAN", ean))
            if (
                upc_as_well
                and upc
                and upc != ean
                and identifier_query_key("UPC", upc) not in skipped_identifier_keys
            ):
                identifiers.append(("UPC", upc))
            if include_title:
                identifiers.append(("TITLE", cleaned_title))
                emitted_title_searches.add(title_key)
            if not identifiers:
                if cleaned_title_as_well and title_key in emitted_title_searches:
                    continue
                if has_identifier_candidate:
                    continue
                identifiers.append(("QUERY", ""))

        for identifier_type, identifier in identifiers:
            variant = product.copy()
            variant["search_identifier_type"] = identifier_type
            variant["search_identifier"] = identifier
            if identifier:
                variant["search_url"] = ebay_search_url_for_identifier(
                    product["search_url"],
                    identifier,
                )
            expanded.append(variant)
    return expanded


def select_products(
    products: list[dict[str, str]],
    settings: ScrapeSettings,
    *,
    skipped_identifier_keys: set[str] | None = None,
) -> tuple[list[dict[str, str]], str]:
    start = settings.start_row
    end = settings.end_row
    limit = settings.limit
    min_sales_rank = settings.min_sales_rank
    max_sales_rank = settings.max_sales_rank
    min_drop_count = settings.min_drop_count
    min_buybox_price = settings.min_buybox_price
    upc_as_well = settings.upc_as_well
    cleaned_title_as_well = settings.cleaned_title_as_well
    titles_only = settings.titles_only

    total_rows = len(products)
    selected = products
    if start is not None:
        selected = [
            product for product in selected if int(product["csv_row"]) >= start
        ]
    if end is not None:
        selected = [
            product for product in selected if int(product["csv_row"]) <= end
        ]
    rows_in_range = len(selected)

    include_brand_keys = {
        brand.strip().casefold()
        for brand in settings.include_brands
        if brand.strip()
    }
    exclude_brand_keys = {
        brand.strip().casefold()
        for brand in settings.exclude_brands
        if brand.strip()
    }
    if include_brand_keys:
        selected = [
            product
            for product in selected
            if product["brand"].strip().casefold() in include_brand_keys
        ]
    if exclude_brand_keys:
        selected = [
            product
            for product in selected
            if product["brand"].strip().casefold() not in exclude_brand_keys
        ]

    if min_drop_count is not None:
        selected = [
            product
            for product in selected
            if (
                (
                    (drops := parse_sales_rank(product.get("drops_count")))
                    is not None
                    and drops >= min_drop_count
                )
                or (
                    drops is None
                    and sales_rank_qualified(
                        product.get("sales_rank"),
                        min_sales_rank=min_sales_rank,
                        max_sales_rank=max_sales_rank,
                    )
                )
            )
        ]
    elif min_sales_rank is not None or max_sales_rank is not None:
        selected = [
            product
            for product in selected
            if sales_rank_qualified(
                product.get("sales_rank"),
                min_sales_rank=min_sales_rank,
                max_sales_rank=max_sales_rank,
            )
        ]

    before_buybox_filter = len(selected)
    selected = [
        product
        for product in selected
        if (
            (buybox := parse_price(product.get("buybox_price"))) is not None
            and (
                min_buybox_price is None
                or buybox >= min_buybox_price
            )
        )
    ]
    skipped_without_buybox = before_buybox_filter - len(selected)

    scrapeable = [product for product in selected if product["search_url"]]
    skipped_without_url = len(selected) - len(scrapeable)
    selected = scrapeable
    skipped_without_title = 0
    if titles_only:
        with_title = [
            product
            for product in selected
            if product.get("cleaned_title", "").strip()
        ]
        skipped_without_title = len(selected) - len(with_title)
        selected = with_title
    filtered_total = len(selected)
    if limit is not None:
        selected = selected[:limit]

    selected = expand_identifier_searches(
        selected,
        upc_as_well=upc_as_well,
        cleaned_title_as_well=cleaned_title_as_well,
        titles_only=titles_only,
        skipped_identifier_keys=skipped_identifier_keys,
    )
    query_total = len(selected)
    for query_position, product in enumerate(selected, start=1):
        product["selection_position"] = query_position
        product["selection_total"] = query_total

    if start is not None or end is not None:
        last_excel_row = max((int(product["csv_row"]) for product in products), default=1)
        range_label = f"Excel rows {start or 2}-{end or last_excel_row}"
    elif limit is not None:
        range_label = f"first {limit} filtered products"
    else:
        range_label = "all filtered products"

    brand_filter_parts = []
    if include_brand_keys:
        brand_filter_parts.append(f"brands: {', '.join(sorted(include_brand_keys))}")
    if exclude_brand_keys:
        brand_filter_parts.append(
            f"excluding brands: {', '.join(sorted(exclude_brand_keys))}"
        )
    if min_drop_count is not None:
        brand_filter_parts.append(f"minimum drops: {min_drop_count:,}")
        if min_sales_rank is not None:
            brand_filter_parts.append(
                f"fallback minimum sales rank: {min_sales_rank:,}"
            )
        if max_sales_rank is not None:
            brand_filter_parts.append(
                f"fallback maximum sales rank: {max_sales_rank:,}"
            )
    else:
        if min_sales_rank is not None:
            brand_filter_parts.append(f"minimum sales rank: {min_sales_rank:,}")
        if max_sales_rank is not None:
            brand_filter_parts.append(f"maximum sales rank: {max_sales_rank:,}")
    if min_buybox_price is not None:
        brand_filter_parts.append(f"minimum buybox: ${min_buybox_price:,.2f}")
    if titles_only:
        brand_filter_parts.append("cleaned-title searches only")
    else:
        if upc_as_well:
            brand_filter_parts.append("EAN + UPC searches")
        if cleaned_title_as_well:
            brand_filter_parts.append("one cleaned-title search per ASIN")
    brand_filter_label = (
        f"; {'; '.join(brand_filter_parts)}" if brand_filter_parts else ""
    )
    details = (
        f"Selected {range_label}{brand_filter_label}: {len(selected)} searches to scrape "
        f"(of {filtered_total} matching products from {rows_in_range} rows in range, "
        f"{total_rows} source rows)"
    )
    if skipped_without_url:
        details += f", skipped {skipped_without_url} without eBay URL"
    if skipped_without_buybox:
        details += (
            f", skipped {skipped_without_buybox} without a qualifying buybox"
        )
    if skipped_without_title:
        details += f", skipped {skipped_without_title} without a cleaned title"

    return selected, details


TITLE_MATCH_IGNORED_TOKENS = {
    "brand",
    "new",
    "sealed",
    "official",
    "authentic",
    "free",
    "shipping",
}


def title_match_tokens(value: object) -> list[str]:
    return [
        token
        for token in re.findall(r"[a-z0-9]+(?:-[a-z0-9]+)*", str(value).casefold())
        if token not in TITLE_MATCH_IGNORED_TOKENS
    ]


def title_listing_matches(cleaned_title: str, listing_title: str) -> bool:
    query_tokens = list(dict.fromkeys(title_match_tokens(cleaned_title)))
    listing_tokens = set(title_match_tokens(listing_title))
    if len(query_tokens) < 2 or not listing_tokens:
        return False

    model_tokens = {
        token
        for token in query_tokens
        if (
            (any(character.isalpha() for character in token)
             and any(character.isdigit() for character in token))
            or (token.isdigit() and len(token) >= 4)
        )
    }
    if model_tokens and model_tokens.isdisjoint(listing_tokens):
        return False

    matched_weight = 0.0
    total_weight = 0.0
    for token in query_tokens:
        weight = 2.0 if any(character.isdigit() for character in token) else 1.0
        total_weight += weight
        if token in listing_tokens:
            matched_weight += weight
    return matched_weight / total_weight >= 0.55


def extract_winners_from_result(
    result: dict,
    *,
    min_roi_percent: float,
    max_listing_age_days: int,
    min_seller_reviews: int = 0,
    max_roi_percent: float | None = None,
) -> list[dict]:
    winners: list[dict] = []
    buybox_value = result.get("buybox_price_value")
    if buybox_value is None:
        buybox_value = parse_price(result.get("buybox_price", ""))

    for listing in filter_us_listings(result.get("listings", [])):
        if (
            result.get("search_identifier_type") == "TITLE"
            and not title_listing_matches(
                str(result.get("cleaned_title") or ""),
                str(listing.get("listing_title") or ""),
            )
        ):
            continue

        evaluated = evaluate_winning_listing(
            listing,
            buybox_price=buybox_value,
            min_roi_percent=min_roi_percent,
            max_listing_age_days=max_listing_age_days,
            min_seller_reviews=min_seller_reviews,
            max_roi_percent=max_roi_percent,
        )
        if not evaluated.get("is_winner"):
            continue

        roi_percent = evaluated.get("roi_percent")
        winners.append(
            {
                "title": result.get("title", ""),
                "ASIN": result.get("asin", ""),
                "EAN": result.get("ean", ""),
                "Brand": result.get("brand", ""),
                "BUYBOX": result.get("buybox_price", ""),
                "AMAZON URL": result.get("amazon_url", ""),
                "EBAY full cost": evaluated.get("spent"),
                "SELLER": evaluated.get("seller_name", ""),
                "SELLER REVIEWS": evaluated.get("seller_reviews_count"),
                "LISTING DATE": (
                    evaluated.get("listing_date_iso")
                    or evaluated.get("date_listed", "")
                ),
                "ROI": None if roi_percent is None else round(roi_percent, 2),
                "EBAY listing URL": evaluated.get("listing_url", ""),
                "ebay listing query": result.get("search_url", ""),
            }
        )

    return winners


def verify_winners_amazon_sales_rank(
    page,
    winners: list[dict],
    *,
    amazon_url: str,
    min_sales_rank: int | None,
    max_sales_rank: int | None,
    cache: dict[str, int | None],
) -> list[dict]:
    if not winners or (min_sales_rank is None and max_sales_rank is None):
        return winners

    amazon_url = amazon_url.strip()
    if page is None or not amazon_url:
        print("  Skipping winners: Amazon sales rank could not be verified")
        return []

    if amazon_url not in cache:
        print("  Verifying Amazon sales rank once for this query")
        try:
            cache[amazon_url] = fetch_amazon_sales_rank(page, amazon_url)
        except Exception as error:
            cache[amazon_url] = None
            print(f"  Amazon lookup failed: {type(error).__name__}: {error}")
    else:
        print("  Using Amazon sales rank already verified for this product")

    sales_rank = cache[amazon_url]
    if sales_rank_qualified(
        sales_rank,
        min_sales_rank=min_sales_rank,
        max_sales_rank=max_sales_rank,
    ):
        print(f"  Amazon sales rank {sales_rank:,} meets criteria")
        return winners

    rank_label = "missing" if sales_rank is None else f"{sales_rank:,}"
    print(
        f"  Skipping {len(winners)} winners: Amazon sales rank {rank_label} "
        "does not meet criteria"
    )
    return []


def _alternate_output_path(output_path: Path, attempt: int) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if attempt == 0:
        return output_path.with_name(f"{output_path.stem}_{stamp}{output_path.suffix}")
    return output_path.with_name(
        f"{output_path.stem}_{stamp}_{attempt}{output_path.suffix}"
    )


def write_winners_xlsx(winners: list[dict], output_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Winning Listings"
    sheet.append(WINNER_HEADERS)

    col_indexes = {header: index + 1 for index, header in enumerate(WINNER_HEADERS)}

    for winner in winners:
        sheet.append([winner.get(header, "") for header in WINNER_HEADERS])

    for row_idx in range(2, len(winners) + 2):
        for header in TEXT_COLUMNS:
            cell = sheet.cell(row=row_idx, column=col_indexes[header])
            cell.value = "" if cell.value is None else str(cell.value)
            cell.number_format = "@"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    candidates = [output_path, *(_alternate_output_path(output_path, i) for i in range(5))]
    last_error: PermissionError | None = None
    for candidate in candidates:
        try:
            workbook.save(candidate)
            return candidate
        except PermissionError as error:
            last_error = error

    raise PermissionError(
        f"Could not save winning listings; file may be open in another program: {output_path}"
    ) from last_error


def write_winners_json(winners: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_suffix(f"{output_path.suffix}.tmp")
    temporary_path.write_text(
        json.dumps(winners, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )
    temporary_path.replace(output_path)


def winner_listing_key(winner: dict) -> str:
    url = str(winner.get("EBAY listing URL") or "").strip()
    if not url:
        return ""

    parts = urlsplit(url)
    item_match = re.search(
        r"/itm/(?:[^/?#]+/)?(\d{9,})",
        parts.path,
        re.IGNORECASE,
    )
    if item_match:
        return f"ebay-item:{item_match.group(1)}"

    normalized_path = parts.path.rstrip("/")
    return urlunsplit(
        (
            parts.scheme.casefold(),
            parts.netloc.casefold(),
            normalized_path,
            "",
            "",
        )
    )


def load_winner_records(path: Path) -> list[dict]:
    if not path.exists():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list) or not all(isinstance(item, dict) for item in data):
        raise ValueError(f"Winner history must contain a JSON list: {path}")
    return data


def initialize_identifier_no_match_history(
    history_path: Path,
    retention_days: int,
) -> tuple[list[dict], set[str], set[str]]:
    history = load_winner_records(history_path)
    retained: list[dict] = []
    active_keys: set[str] = set()
    pending_deletions: set[str] = set()
    seen_keys: set[str] = set()
    now = datetime.now()

    for record in history:
        key = identifier_query_key(
            str(record.get("query_type") or ""),
            record.get("identifier"),
        )
        timestamp = str(record.get("no_exact_match_at") or "")
        if not key or not timestamp:
            continue
        try:
            recorded_at = datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
            if recorded_at.tzinfo is not None:
                recorded_at = recorded_at.replace(tzinfo=None)
        except ValueError:
            continue
        if key in seen_keys:
            continue
        seen_keys.add(key)

        stored = record.copy()
        stored["consecutive_no_matches"] = max(
            1,
            int(record.get("consecutive_no_matches") or 1),
        )
        retained.append(stored)
        if stored["consecutive_no_matches"] >= 2:
            pending_deletions.add(key)
            active_keys.add(key)
        elif now < recorded_at + timedelta(days=retention_days):
            active_keys.add(key)

    if retained != history:
        write_winners_json(retained, history_path)
    return retained, active_keys, pending_deletions


def record_identifier_no_match(
    product: dict[str, str],
    *,
    history: list[dict],
    history_keys: set[str],
    history_path: Path,
    retention_days: int,
) -> bool:
    query_type = str(product.get("search_identifier_type") or "").upper()
    identifier = str(product.get("search_identifier") or "").strip()
    if query_type not in {"EAN", "UPC"}:
        return False
    key = identifier_query_key(query_type, identifier)
    if not key:
        return False

    history_keys.add(key)
    existing = next(
        (
            record
            for record in history
            if identifier_query_key(
                str(record.get("query_type") or ""),
                record.get("identifier"),
            )
            == key
        ),
        None,
    )
    if existing is None:
        existing = {
            "query_type": query_type,
            "identifier": identifier,
            "consecutive_no_matches": 0,
        }
        history.append(existing)

    consecutive = int(existing.get("consecutive_no_matches") or 0) + 1
    existing["consecutive_no_matches"] = consecutive
    existing["no_exact_match_at"] = datetime.now().isoformat(timespec="seconds")
    write_winners_json(history, history_path)
    if consecutive >= 2:
        print(
            f"  {query_type} {identifier} returned no exact matches twice; "
            "scheduled for deletion"
        )
        return True

    print(
        f"  Cached {query_type} {identifier}: no exact matches for "
        f"{retention_days} days"
    )
    return False


def clear_identifier_no_match_streak(
    product: dict[str, str],
    *,
    history: list[dict],
    history_keys: set[str],
    history_path: Path,
) -> None:
    query_type = str(product.get("search_identifier_type") or "").upper()
    identifier = str(product.get("search_identifier") or "").strip()
    key = identifier_query_key(query_type, identifier)
    if query_type not in {"EAN", "UPC"} or not key:
        return

    retained = [
        record
        for record in history
        if identifier_query_key(
            str(record.get("query_type") or ""),
            record.get("identifier"),
        )
        != key
    ]
    if len(retained) == len(history):
        return
    history[:] = retained
    history_keys.discard(key)
    write_winners_json(history, history_path)


def delete_identifiers_from_workbook(
    path: Path,
    deletion_keys: set[str],
) -> int:
    if not deletion_keys or path.suffix.casefold() != ".xlsx":
        return 0

    workbook = load_workbook(path)
    worksheet = workbook["main"] if "main" in workbook.sheetnames else workbook.active
    columns = {
        str(cell.value or "").strip().upper(): cell.column
        for cell in worksheet[1]
    }
    missing = {"EAN", "UPC"}.difference(columns)
    if missing:
        workbook.close()
        raise ValueError(
            f"Cannot delete identifiers; missing columns: {', '.join(sorted(missing))}"
        )

    changed = 0
    for row_number in range(2, worksheet.max_row + 1):
        for query_type in ("EAN", "UPC"):
            cell = worksheet.cell(
                row=row_number,
                column=columns[query_type],
            )
            codes = [
                code.strip()
                for code in str(cell.value or "").split(",")
                if code.strip()
            ]
            remaining = [
                code
                for code in codes
                if identifier_query_key(query_type, code) not in deletion_keys
            ]
            if remaining == codes:
                continue
            cell.value = ", ".join(remaining)
            cell.number_format = "@"
            changed += 1

    backup_path = create_workbook_backup(path)
    temporary_path = path.with_name(f".{path.stem}.tmp{path.suffix}")
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        workbook.close()
        temporary_path.unlink(missing_ok=True)
    return changed


def apply_identifier_deletions(
    input_path: Path,
    deletion_keys: set[str],
    *,
    history: list[dict],
    history_keys: set[str],
    history_path: Path,
) -> int:
    if not deletion_keys:
        return 0
    changed = delete_identifiers_from_workbook(input_path, deletion_keys)
    history[:] = [
        record
        for record in history
        if identifier_query_key(
            str(record.get("query_type") or ""),
            record.get("identifier"),
        )
        not in deletion_keys
    ]
    history_keys.difference_update(deletion_keys)
    write_winners_json(history, history_path)
    deletion_keys.clear()
    return changed


def initialize_winner_history(
    history_path: Path,
    live_json_path: Path,
    retention_days: int | None,
) -> tuple[list[dict], set[str], int, int]:
    history_exists = history_path.exists()
    history = load_winner_records(history_path)
    history_keys: set[str] = set()
    deduplicated_history: list[dict] = []
    pruned = 0
    now = datetime.now()
    cutoff = (
        now - timedelta(days=retention_days)
        if retention_days is not None
        else None
    )

    for record in history:
        first_won_at = str(record.get("first_won_at") or "")
        if cutoff is not None and first_won_at:
            try:
                won_at = datetime.fromisoformat(first_won_at.replace("Z", "+00:00"))
                if won_at.tzinfo is not None:
                    won_at = won_at.replace(tzinfo=None)
                if won_at < cutoff:
                    pruned += 1
                    continue
            except ValueError:
                pass

        key = str(record.get("history_key") or winner_listing_key(record))
        if not key or key in history_keys:
            continue
        history_keys.add(key)
        stored = record.copy()
        stored["history_key"] = key
        if not first_won_at:
            stored["first_won_at"] = now.isoformat(timespec="seconds")
        deduplicated_history.append(stored)

    seeded = 0
    if not history_exists and live_json_path.exists():
        for winner in load_winner_records(live_json_path):
            key = winner_listing_key(winner)
            if not key or key in history_keys:
                continue
            history_keys.add(key)
            stored = winner.copy()
            stored["history_key"] = key
            stored["first_won_at"] = now.isoformat(timespec="seconds")
            deduplicated_history.append(stored)
            seeded += 1

    if seeded or deduplicated_history != history:
        write_winners_json(deduplicated_history, history_path)

    return deduplicated_history, history_keys, seeded, pruned


def persist_new_winners(
    candidates: list[dict],
    *,
    winners: list[dict],
    session_keys: set[str],
    history: list[dict],
    history_keys: set[str],
    skip_previously_won: bool,
    live_json_path: Path,
    history_path: Path,
) -> tuple[int, int]:
    accepted: list[dict] = []
    history_changed = False
    skipped = 0

    for winner in candidates:
        key = winner_listing_key(winner)
        if key and key in session_keys:
            skipped += 1
            continue
        if key and skip_previously_won and key in history_keys:
            skipped += 1
            continue

        accepted.append(winner)
        if key:
            session_keys.add(key)
        if key and key not in history_keys:
            history_keys.add(key)
            stored = winner.copy()
            stored["history_key"] = key
            stored["first_won_at"] = datetime.now().isoformat(timespec="seconds")
            history.append(stored)
            history_changed = True

    if history_changed:
        write_winners_json(history, history_path)
    if accepted:
        winners.extend(accepted)
        write_winners_json(winners, live_json_path)
    return len(accepted), skipped


def query_label(product: dict[str, str]) -> tuple[str, str]:
    query_type = str(product.get("search_identifier_type") or "EAN")
    identifier = (
        product.get("search_identifier")
        or product.get("ean")
        or product.get("asin")
        or "(unknown)"
    )
    return query_type, str(identifier)


def scrape_product_result(
    page,
    product: dict[str, str],
    *,
    saved_html: str | None,
) -> dict:
    common_kwargs = {
        "title": product["title"],
        "asin": product["asin"],
        "ean": product["ean"],
        "buybox_price": product["buybox_price"],
    }

    if saved_html is not None:
        result = scrape_search_page_from_html(
            product["search_url"],
            saved_html,
            **common_kwargs,
        )
    else:
        result = scrape_search_page(
            page,
            product["search_url"],
            **common_kwargs,
        )

    result["amazon_url"] = product["amazon_url"]
    result["brand"] = product["brand"]
    result["search_identifier_type"] = product.get("search_identifier_type", "")
    result["cleaned_title"] = product.get("cleaned_title", "")
    return result


def process_product(
    page,
    product: dict[str, str],
    *,
    index: int,
    total: int,
    saved_html: str | None,
    settings: ScrapeSettings,
    no_exact_match_callback: Callable[[dict[str, str]], None] | None = None,
    exact_match_callback: Callable[[dict[str, str]], None] | None = None,
    seller_review_cache: dict[str, tuple[str, int | None]] | None = None,
    amazon_rank_cache: dict[str, int | None] | None = None,
) -> list[dict]:
    identifier_type, identifier = query_label(product)
    result = scrape_product_result(page, product, saved_html=saved_html)
    query_type = product.get("search_identifier_type")
    if query_type in {"EAN", "UPC"}:
        if not result.get("listings"):
            if no_exact_match_callback is not None:
                no_exact_match_callback(product)
        elif exact_match_callback is not None:
            exact_match_callback(product)

    us_listings = filter_us_listings(result.get("listings", []))
    winner_kw = settings.winner_kwargs()
    fill_missing_seller_reviews(
        None if saved_html is not None else page,
        us_listings,
        buybox_price=result.get("buybox_price_value"),
        cache=seller_review_cache,
        **winner_kw,
    )
    product_winners = extract_winners_from_result(result, **winner_kw)
    product_winners = verify_winners_amazon_sales_rank(
        None if saved_html is not None else page,
        product_winners,
        amazon_url=str(product.get("amazon_url") or result.get("amazon_url") or ""),
        min_sales_rank=settings.min_sales_rank,
        max_sales_rank=settings.max_sales_rank,
        cache=amazon_rank_cache if amazon_rank_cache is not None else {},
    )

    winner_status = f"yes ({len(product_winners)})" if product_winners else "no"
    print(f"[{index}/{total}] {identifier_type} {identifier} - winners: {winner_status}")
    print(f"  US listings found: {len(us_listings)}")
    print(f"  {format_cheapest_listing_log(result, **winner_kw)}")
    for line in format_below_buybox_listing_logs(result, **winner_kw):
        print(line)
    block_log = format_block_log(result)
    if block_log:
        print(f"  {block_log}")
    return product_winners


def process_live_product_with_ship_to_retry(
    page,
    product: dict[str, str],
    **kwargs,
) -> list[dict]:
    try:
        return process_product(page, product, saved_html=None, **kwargs)
    except EbayShipToNotUsError as error:
        identifier_type, identifier = query_label(product)
        print(f"  Ship to is not US for {identifier_type} {identifier}: {error}")
        print("  Refreshing and retrying this search")
        refresh_and_verify_ship_to_us(page)
        return process_product(page, product, saved_html=None, **kwargs)


def main(settings: ScrapeSettings | None = None) -> int:
    settings = settings or ScrapeSettings()
    try:
        settings.validate()
    except ValueError as error:
        print(f"Error: {error}")
        return 1

    input_file = settings.input_csv
    output_file = settings.output_xlsx
    json_output_file = settings.live_json
    history_file = settings.winner_history
    identifier_no_match_file = settings.identifier_no_match_history

    ensure_data_dirs()
    if not input_file.exists():
        print(f"Error: input file not found: {input_file}")
        return 1

    deleted_backups = prune_directory_workbook_backups(input_file.parent)
    if deleted_backups:
        print(f"Deleted {len(deleted_backups)} old workbook backup(s)")

    try:
        (
            identifier_no_match_history,
            identifier_no_match_keys,
            pending_identifier_deletions,
        ) = initialize_identifier_no_match_history(
            identifier_no_match_file,
            settings.identifier_no_match_retention_days,
        )
    except Exception as error:
        print(f"Error loading identifier no-match history: {error}")
        return 1
    if pending_identifier_deletions:
        try:
            deleted_cells = apply_identifier_deletions(
                input_file,
                pending_identifier_deletions,
                history=identifier_no_match_history,
                history_keys=identifier_no_match_keys,
                history_path=identifier_no_match_file,
            )
            print(
                f"Deleted {deleted_cells} EAN/UPC cells after repeated "
                "no-match results"
            )
        except Exception as error:
            print(f"Error applying pending identifier deletions: {error}")
            return 1
    if identifier_no_match_keys:
        print(
            f"Skipping {len(identifier_no_match_keys)} EAN/UPC queries "
            "with recent no-match results"
        )

    try:
        products, selection_details = select_products(
            load_products(input_file),
            settings,
            skipped_identifier_keys=identifier_no_match_keys,
        )
    except ValueError as error:
        print(f"Error: {error}")
        return 1

    if not products:
        print(f"Error: no rows selected to scrape from {input_file}")
        return 1

    print(selection_details)

    winners: list[dict] = []
    session_winner_keys: set[str] = set()
    seller_review_cache: dict[str, tuple[str, int | None]] = {}
    amazon_rank_cache: dict[str, int | None] = {}
    total = len(products)
    exit_code = 0
    skipped_previous_winners = 0
    current_product: dict[str, str] | None = None
    winner_history: list[dict] = []
    winner_history_keys: set[str] = set()

    def log_current_stop(reason: str) -> None:
        product = current_product or (products[0] if products else None)
        log_workbook_stop(
            workbook_path=input_file,
            excel_row=None if product is None else product.get("csv_row"),
            reason=reason,
            script="scrape",
            extra=scrape_stop_extra(product),
        )

    def cache_identifier_no_match(product: dict[str, str]) -> None:
        should_delete = record_identifier_no_match(
            product,
            history=identifier_no_match_history,
            history_keys=identifier_no_match_keys,
            history_path=identifier_no_match_file,
            retention_days=settings.identifier_no_match_retention_days,
        )
        if should_delete:
            pending_identifier_deletions.add(
                identifier_query_key(
                    str(product.get("search_identifier_type") or ""),
                    product.get("search_identifier"),
                )
            )

    def clear_identifier_streak(product: dict[str, str]) -> None:
        clear_identifier_no_match_streak(
            product,
            history=identifier_no_match_history,
            history_keys=identifier_no_match_keys,
            history_path=identifier_no_match_file,
        )

    def keep_winners(product_winners: list[dict]) -> None:
        nonlocal skipped_previous_winners
        if not product_winners:
            return
        _accepted, skipped = persist_new_winners(
            product_winners,
            winners=winners,
            session_keys=session_winner_keys,
            history=winner_history,
            history_keys=winner_history_keys,
            skip_previously_won=settings.skip_previously_won,
            live_json_path=json_output_file,
            history_path=history_file,
        )
        skipped_previous_winners += skipped

    def process_kwargs(index: int, total_count: int, *, live: bool = False) -> dict:
        kwargs = {
            "index": index,
            "total": total_count,
            "settings": settings,
            "seller_review_cache": seller_review_cache,
            "amazon_rank_cache": amazon_rank_cache,
        }
        if live:
            kwargs["no_exact_match_callback"] = cache_identifier_no_match
            kwargs["exact_match_callback"] = clear_identifier_streak
        return kwargs

    try:
        winner_history, winner_history_keys, seeded_history, pruned_history = (
            initialize_winner_history(
                history_file,
                json_output_file,
                settings.winner_history_retention_days,
            )
        )
    except Exception as error:
        print(f"Error loading winner history: {error}")
        return 1

    if seeded_history:
        print(
            f"Seeded winner history with {seeded_history} listings "
            "from the previous live JSON"
        )
    if pruned_history:
        print(
            f"Deleted {pruned_history} winner-history entries older than "
            f"{settings.winner_history_retention_days} days"
        )
    print(
        f"Winner history: {len(winner_history_keys)} listings "
        f"({'skipping matches' if settings.skip_previously_won else 'tracking only'})"
    )

    try:
        write_winners_json(winners, json_output_file)
        saved_html = (
            settings.html_path.read_text(encoding="utf-8")
            if settings.html_path
            else None
        )
        if saved_html is not None:
            print("Offline test mode: parsing saved HTML (Playwright not used)")
            for index, product in enumerate(products, start=1):
                current_product = product
                display_index = int(product.get("selection_position", index))
                display_total = int(product.get("selection_total", total))
                keep_winners(
                    process_product(
                        None,
                        product,
                        saved_html=saved_html,
                        **process_kwargs(display_index, display_total),
                    )
                )
        else:
            cookies = load_ebay_cookies(
                cookie_header=settings.cookie_header,
                cookies_file=settings.cookies_file,
                default_cookies_file=EBAY_COOKIES_FILE,
            )
            print(f"Scraping {total} eBay search URLs with Playwright")
            print(f"eBay session: {describe_ebay_cookie_session(cookies)}")
            if BROWSER_RESTART_EVERY:
                print(f"Restarting browser every {BROWSER_RESTART_EVERY} searches")
            with browser_session(cookies=cookies) as session:
                skipped_previous_ean_for_ship_to = False
                searches_since_browser_start = 0
                for index, product in enumerate(products, start=1):
                    current_product = product
                    if (
                        BROWSER_RESTART_EVERY
                        and searches_since_browser_start >= BROWSER_RESTART_EVERY
                    ):
                        print(
                            f"Restarting browser after {index - 1} searches "
                            "to free memory"
                        )
                        session.restart()
                        searches_since_browser_start = 0
                    searches_since_browser_start += 1
                    display_index = int(product.get("selection_position", index))
                    display_total = int(product.get("selection_total", total))
                    query_type = str(
                        product.get("search_identifier_type") or ""
                    ).upper()
                    query_key = identifier_query_key(
                        query_type,
                        product.get("search_identifier"),
                    )
                    if (
                        query_type in {"EAN", "UPC"}
                        and query_key in identifier_no_match_keys
                    ):
                        print(
                            f"[{display_index}/{display_total}] Skipping "
                            f"{query_type} {product.get('search_identifier')}: "
                            "recently had no exact matches"
                        )
                        continue
                    try:
                        product_winners = process_live_product_with_ship_to_retry(
                            session.page,
                            product,
                            **process_kwargs(display_index, display_total, live=True),
                        )
                    except EbayShipToNotUsError as error:
                        ean = product.get("ean") or product.get("asin") or "(unknown)"
                        if skipped_previous_ean_for_ship_to:
                            raise EbayShipToNotUsError(
                                "consecutive_not_us",
                                detail=(
                                    f"EAN {ean} was still not US after refresh; "
                                    "the previous EAN was also skipped"
                                ),
                            ) from error
                        print(
                            f"  Skipping EAN {ean}: Ship to is still not US "
                            "after refresh"
                        )
                        skipped_previous_ean_for_ship_to = True
                        continue

                    skipped_previous_ean_for_ship_to = False
                    keep_winners(product_winners)
    except EbayShipToNotUsError as error:
        print(f"Stopping scrape: {error}")
        log_current_stop(str(error))
        exit_code = 1
    except KeyboardInterrupt:
        print("Scrape stopped by user; saving winners collected so far")
        log_current_stop("stopped by user")
        exit_code = 130
    except Exception as error:
        print(f"Scrape stopped by {type(error).__name__}: {error}")
        log_current_stop(f"{type(error).__name__}: {error}")
        exit_code = 1

    if pending_identifier_deletions:
        try:
            deleted_cells = apply_identifier_deletions(
                input_file,
                pending_identifier_deletions,
                history=identifier_no_match_history,
                history_keys=identifier_no_match_keys,
                history_path=identifier_no_match_file,
            )
            print(
                f"Deleted {deleted_cells} EAN/UPC cells after two "
                "consecutive no-match results"
            )
        except Exception as error:
            print(f"Error deleting repeated no-match identifiers: {error}")
            exit_code = 1

    try:
        write_winners_json(winners, json_output_file)
        print(f"Live JSON feed saved to {json_output_file}")
    except Exception as error:
        print(f"Error saving live JSON feed: {error}")
        exit_code = 1

    try:
        saved_path = write_winners_xlsx(winners, output_file)
    except Exception as error:
        print(f"Error saving winning listings: {error}")
        return 1

    print(f"Found {len(winners)} winning listings")
    if skipped_previous_winners:
        print(
            f"Skipped {skipped_previous_winners} previously won "
            "or duplicate listings"
        )
    if saved_path != output_file:
        print(
            f"Could not overwrite {output_file.name} (file may be open); "
            f"saved to {saved_path.name}"
        )
    else:
        print(f"Saved to {saved_path}")
    return exit_code


if __name__ == "__main__":
    print("Edit SETTINGS in ebay/main.py, then run that file.")
    raise SystemExit(0)

