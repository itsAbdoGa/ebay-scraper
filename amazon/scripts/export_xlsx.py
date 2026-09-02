import argparse
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook, load_workbook

from lib.paths import PRODUCT_DETAILS_JSON, PRODUCT_DETAILS_XLSX, ensure_data_dirs

BRAND_KEY = "brand"
CATEGORY_KEY = "main_category"
BRAND_PREFIX = "Brand: "


def clean_brand_name(brand: str) -> str:
    brand = brand.strip()
    if brand.startswith(BRAND_PREFIX):
        return brand[len(BRAND_PREFIX) :].strip()
    return brand


def normalize_brands(records: list[dict]) -> list[dict]:
    normalized: list[dict] = []
    for record in records:
        updated = dict(record)
        updated[BRAND_KEY] = clean_brand_name(str(record.get(BRAND_KEY, "")))
        normalized.append(updated)
    return normalized


def deduplicate_by_brand_category(records: list[dict]) -> list[dict]:
    seen: set[tuple[str, str]] = set()
    deduped: list[dict] = []

    for record in records:
        key = (record.get(BRAND_KEY, ""), record.get(CATEGORY_KEY, ""))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(record)

    return deduped


def prepare_export_records(
    records: list[dict],
    *,
    deduplicate: bool = True,
    sort_by_category: bool = True,
) -> list[dict]:
    records = normalize_brands(records)
    if deduplicate:
        records = deduplicate_by_brand_category(records)
    if sort_by_category:
        records = sorted(
            records,
            key=lambda record: (
                str(record.get(CATEGORY_KEY, "")).casefold(),
                str(record.get(BRAND_KEY, "")).casefold(),
            ),
        )
    return records


def load_records(path: Path) -> list[dict]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError(f"{path} must contain a JSON array")

    records: list[dict] = []
    for index, item in enumerate(data, start=1):
        if not isinstance(item, dict):
            raise ValueError(f"{path} item {index} must be a JSON object")
        records.append(item)

    return records


def load_xlsx_records(path: Path) -> list[dict]:
    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(header) for header in rows[0]]
    records: list[dict] = []

    for row in rows[1:]:
        record = {
            header: "" if index >= len(row) or row[index] is None else row[index]
            for index, header in enumerate(headers)
        }
        records.append(record)

    return records


def collect_headers(records: list[dict]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()

    for record in records:
        for key in record:
            if key not in seen:
                seen.add(key)
                headers.append(key)

    return headers


def write_xlsx(records: list[dict], output_path: Path) -> None:
    headers = collect_headers(records)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"

    sheet.append(headers)

    for record in records:
        sheet.append([record.get(header, "") for header in headers])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert product details JSON to Excel (xlsx)."
    )
    parser.add_argument(
        "-i",
        "--input",
        type=Path,
        default=PRODUCT_DETAILS_JSON,
        help=f"Input JSON file (default: {PRODUCT_DETAILS_JSON})",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=PRODUCT_DETAILS_XLSX,
        help=f"Output xlsx file (default: {PRODUCT_DETAILS_XLSX})",
    )
    parser.add_argument(
        "--no-deduplicate",
        action="store_true",
        help="Keep all rows instead of one row per brand/category pair",
    )
    return parser.parse_args()


def main(
    input_path: Path | None = None,
    output_path: Path | None = None,
    deduplicate: bool = True,
) -> int:
    input_file = input_path or PRODUCT_DETAILS_JSON
    output_file = output_path or PRODUCT_DETAILS_XLSX

    ensure_data_dirs()

    if not input_file.exists():
        print(f"Error: product details file not found: {input_file}")
        print("Run 'python run.py scrape' first.")
        return 1

    records = load_records(input_file)
    original_count = len(records)
    records = prepare_export_records(records, deduplicate=deduplicate)

    write_xlsx(records, output_file)

    print(f"Read {original_count} records from {input_file}")
    if deduplicate:
        removed = original_count - len(records)
        print(f"Deduplicated to {len(records)} unique brand/category rows ({removed} removed)")
    print(f"Wrote {output_file}")
    return 0


if __name__ == "__main__":
    args = parse_args()
    raise SystemExit(
        main(
            input_path=args.input,
            output_path=args.output,
            deduplicate=not args.no_deduplicate,
        )
    )
