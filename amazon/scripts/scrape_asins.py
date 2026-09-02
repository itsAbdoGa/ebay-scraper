import argparse
import csv
import json
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from playwright.sync_api import sync_playwright

from lib.amazon_scraper import (
    REQUEST_DELAY_SECONDS,
    create_browser_context,
    scrape_product,
    warm_up_session,
)
from lib.paths import (
    ALL_ASIN_DETAILS_DEDUPED_JSON,
    ALL_ASIN_DETAILS_JSON,
    ALL_ASIN_DETAILS_XLSX,
    ALL_ASINS_CSV,
    ALREADY_REGISTERED_CSV,
    ensure_data_dirs,
)
from scripts.export_xlsx import clean_brand_name, prepare_export_records, write_xlsx

ASIN_RE = re.compile(r"^[A-Z0-9]{10}$")
ASIN_FROM_URL_RE = re.compile(r"/dp/([A-Z0-9]{10})", re.IGNORECASE)


@dataclass
class RegisteredRegistry:
    asins: set[str]
    brand_categories: set[tuple[str, str]]


def normalize_brand_category_key(brand: str, category: str) -> tuple[str, str]:
    return (clean_brand_name(brand).casefold(), category.strip().casefold())


def load_registered_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        rows: list[dict[str, str]] = []
        try:
            for worksheet in workbook.worksheets:
                values = worksheet.iter_rows(values_only=True)
                headers = next(values, None)
                if not headers:
                    continue

                header_names = [str(value).strip() if value is not None else "" for value in headers]
                for values_row in values:
                    rows.append(
                        {
                            header: str(value).strip() if value is not None else ""
                            for header, value in zip(header_names, values_row)
                            if header
                        }
                    )
        finally:
            workbook.close()
        return rows

    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader) if reader.fieldnames else []


def load_registered_registry(path: Path) -> RegisteredRegistry:
    registered_asins: set[str] = set()
    brand_categories: set[tuple[str, str]] = set()

    for row in load_registered_rows(path):
        product_url = (row.get("Product") or row.get("Link") or "").strip()
        match = ASIN_FROM_URL_RE.search(product_url)
        if match:
            registered_asins.add(match.group(1).upper())

        brand = (row.get("Brand") or "").strip()
        category = (row.get("Category") or "").strip()
        if brand and category:
            brand_categories.add(normalize_brand_category_key(brand, category))

    return RegisteredRegistry(asins=registered_asins, brand_categories=brand_categories)


def is_registered_asin(asin: str, registry: RegisteredRegistry) -> bool:
    return asin.upper() in registry.asins


def is_registered_brand_category(
    brand: str,
    category: str,
    registry: RegisteredRegistry,
) -> bool:
    if not brand.strip() or not category.strip():
        return False
    return normalize_brand_category_key(brand, category) in registry.brand_categories


def load_asins_from_csv(
    path: Path,
    *,
    registry: RegisteredRegistry | None = None,
) -> tuple[list[dict[str, str]], int]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError(f"{path} is empty or missing headers")

        asin_column = None
        title_column = None
        for field in reader.fieldnames:
            normalized = field.strip().lower()
            if normalized == "asin":
                asin_column = field
            if normalized == "title":
                title_column = field

        if asin_column is None:
            raise ValueError(f"{path} must contain an ASIN column")

        listings: list[dict[str, str]] = []
        seen_asins: set[str] = set()
        skipped_registered = 0

        for row_number, row in enumerate(reader, start=2):
            asin = row.get(asin_column, "").strip().upper()
            if not asin:
                continue
            if not ASIN_RE.fullmatch(asin):
                raise ValueError(f"{path}:{row_number} has invalid ASIN: {asin}")
            if asin in seen_asins:
                continue
            if registry is not None and is_registered_asin(asin, registry):
                skipped_registered += 1
                continue

            seen_asins.add(asin)
            title = row.get(title_column or "", "").strip() if title_column else ""
            listings.append(
                {
                    "asin": asin,
                    "title": title,
                    "url": f"https://www.amazon.com/dp/{asin}",
                }
            )

    return listings, skipped_registered


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape brand and category for every ASIN in a CSV file."
    )
    parser.add_argument(
        "-i",
        "--input",
        type=Path,
        default=ALL_ASINS_CSV,
        help=f"CSV file with ASIN column (default: {ALL_ASINS_CSV})",
    )
    parser.add_argument(
        "--registered",
        type=Path,
        default=ALREADY_REGISTERED_CSV,
        help=f"Already registered products CSV (default: {ALREADY_REGISTERED_CSV})",
    )
    parser.add_argument(
        "--json-output",
        type=Path,
        default=ALL_ASIN_DETAILS_JSON,
        help=f"Full scrape results JSON (default: {ALL_ASIN_DETAILS_JSON})",
    )
    parser.add_argument(
        "--deduped-json-output",
        type=Path,
        default=ALL_ASIN_DETAILS_DEDUPED_JSON,
        help=f"Deduplicated JSON (default: {ALL_ASIN_DETAILS_DEDUPED_JSON})",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=ALL_ASIN_DETAILS_XLSX,
        help=f"Deduplicated Excel output (default: {ALL_ASIN_DETAILS_XLSX})",
    )
    parser.add_argument(
        "-n",
        "--limit",
        type=int,
        default=None,
        help="Only scrape the first N ASINs (useful for testing)",
    )
    return parser.parse_args()


def main(
    input_path: Path | None = None,
    registered_path: Path | None = None,
    json_output_path: Path | None = None,
    deduped_json_output_path: Path | None = None,
    xlsx_output_path: Path | None = None,
    limit: int | None = None,
) -> int:
    csv_file = input_path or ALL_ASINS_CSV
    registered_file = registered_path or ALREADY_REGISTERED_CSV
    full_json_file = json_output_path or ALL_ASIN_DETAILS_JSON
    deduped_json_file = deduped_json_output_path or ALL_ASIN_DETAILS_DEDUPED_JSON
    xlsx_file = xlsx_output_path or ALL_ASIN_DETAILS_XLSX

    ensure_data_dirs()

    if not csv_file.exists():
        print(f"Error: CSV file not found: {csv_file}")
        print(f"Place your ASIN export at: {ALL_ASINS_CSV}")
        return 1

    registry = RegisteredRegistry(asins=set(), brand_categories=set())
    if registered_file.exists():
        registry = load_registered_registry(registered_file)
        print(
            f"Loaded {len(registry.asins)} registered ASINs and "
            f"{len(registry.brand_categories)} registered brand/category pairs "
            f"from {registered_file}"
        )
    else:
        print(f"Warning: registered file not found: {registered_file}")

    listings, skipped_asins = load_asins_from_csv(csv_file, registry=registry)
    if limit is not None:
        listings = listings[:limit]

    print(f"Loaded {len(listings)} ASINs from {csv_file}")
    if skipped_asins:
        print(f"Skipped {skipped_asins} already registered ASINs before scraping")
    print()

    results: list[dict[str, str]] = []
    skipped_brand_categories = 0

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=True,
            args=["--disable-blink-features=AutomationControlled"],
        )
        context = create_browser_context(browser)
        page = context.new_page()
        warm_up_session(page)

        try:
            for index, listing in enumerate(listings, start=1):
                label = listing["title"] or listing["url"]
                print(f"[{index}/{len(listings)}] {label}")
                result = scrape_product(
                    page,
                    listing["url"],
                    title=listing["title"],
                )

                brand = clean_brand_name(result.get("brand", ""))
                category = result.get("main_category", "").strip()
                if is_registered_brand_category(brand, category, registry):
                    skipped_brand_categories += 1
                    print(
                        f"  skipped: registered brand/category "
                        f"({brand or '(unknown)'} / {category or '(unknown)'})"
                    )
                else:
                    results.append(result)

                if index < len(listings):
                    time.sleep(REQUEST_DELAY_SECONDS)
        finally:
            context.close()
            browser.close()

    export_records = prepare_export_records(results)

    full_json_file.parent.mkdir(parents=True, exist_ok=True)
    full_json_file.write_text(json.dumps(results, indent=2), encoding="utf-8")
    deduped_json_file.write_text(json.dumps(export_records, indent=2), encoding="utf-8")
    write_xlsx(export_records, xlsx_file)

    print()
    print(f"Wrote {len(results)} scraped records to {full_json_file}")
    if skipped_brand_categories:
        print(
            f"Skipped {skipped_brand_categories} scraped products with "
            f"already registered brand/category pairs"
        )
    print(
        f"Deduplicated to {len(export_records)} unique brand/category rows "
        f"({len(results) - len(export_records)} removed)"
    )
    print(f"Wrote deduplicated JSON to {deduped_json_file}")
    print(f"Wrote deduplicated Excel to {xlsx_file}")
    return 0


if __name__ == "__main__":
    args = parse_args()
    raise SystemExit(
        main(
            input_path=args.input,
            registered_path=args.registered,
            json_output_path=args.json_output,
            deduped_json_output_path=args.deduped_json_output,
            xlsx_output_path=args.output,
            limit=args.limit,
        )
    )
